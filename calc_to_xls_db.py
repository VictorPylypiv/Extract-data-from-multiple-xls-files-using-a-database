import sqlite3
from os.path import join, abspath

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

from def_xls_db import DB

COLUMN_LIST = ['A', 'B', 'C', 'E', 'G', 'H', 'I', 'K']
WIDTH_LIST = [28, 18, 18, 22, 22, 18, 12, 12]


def col_width(ws, column_list, width_list):
    """
    Set the width of the columns on the sheet
    """
    if len(column_list) == len(width_list):
        for i in range(len(column_list)):
            ws.column_dimensions[column_list[i]].width = width_list[i]


def empty_sheet(list, wb, t_header, cur):
    """
    Return a sheet of details without data based on 'empty list'
    """
    if list:
        ws2 = wb.create_sheet(title='empty')
        ws2['A1'] = 'Without calculation'
        d_l = [t_header[1:4]]
        for d in list:
            d_l.append(cur.execute("SELECT ord, det, qty FROM det_data WHERE ord_det = ?", (d,)).fetchone())
        for d in d_l:
            ws2.append(d)

        col_width(ws2, COLUMN_LIST, WIDTH_LIST)


def table_header(cur):
    """
    Create a table header.
    """
    cur.row_factory = lambda cursor, row: row[0]
    t_h_1 = cur.execute("SELECT name FROM PRAGMA_table_info('det_data')").fetchall()
    t_h_2 = cur.execute("SELECT name FROM PRAGMA_table_info('calculation')").fetchall()
    cur.row_factory = None

    t_header = t_h_1 + t_h_2[1:]
    return t_header


def calculation_list(cur):
    """
    Return a list of details with calculations data based on the 'calculation' table.
    Create a table header.
    """
    # conn = sqlite3.connect('./calculations.db')
    # cur = conn.cursor()

    cur.execute("SELECT * FROM det_data d JOIN calculation c on d.ord_det = c.ord_det ORDER BY d.ord_det")
    data_1 = cur.fetchall()
    t_header = table_header(cur)
    data = [t_header]
    # print(data)
    for i in data_1:
        data.append(i[0:4] + i[5:12])
    # print(data)
    # conn.commit()
    # conn.close()

    return data, t_header


def calculation_data(date, e_list=None):
    """
    Creating a file filled with calculation data
    """
    wb = Workbook()
    ws = wb.active
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    data, t_header = calculation_list(cur)
    for row in data:
        ws.append(row)

    if e_list:
        empty_sheet(e_list, wb, t_header, cur)

    conn.commit()
    conn.close()

    col_width(ws, COLUMN_LIST, WIDTH_LIST)

    dxf = DifferentialStyle(fill=PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00'))
    rule = Rule(operator='equal', type='containsText', text='halffabricat', dxf=dxf)
    ws.conditional_formatting.add("H1:H10000", rule)

    new_file = f'calculations_{date}.xlsx'
    new_file_name = abspath(join('.', 'calculation_data', new_file))
    wb.save(new_file_name)
    wb.close()
