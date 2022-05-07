import os
import re
from os.path import join, abspath
from sqlite3 import OperationalError, IntegrityError

from openpyxl import load_workbook

DB = './calculations.db'
# DB = ':memory:'
PATH_C = "\\calculations\\"
PATH_HF = "\\halffabricat_orders\\"
SHEET_NAME = "матеріали"


def create_tables(cur):
    try:
        with open('create_tables.sql', 'r') as sql_file:
            cur.executescript(sql_file.read())
    except OperationalError as msg:
        print("Error: ", msg)


def imp_table_data(cur, ws_table, min_row=2, min_col=1):
    """
    Import data from a sheet table where the cells in the first column are row IDs
    """
    for row in ws_table.iter_rows(min_row=min_row, min_col=min_col, max_row=ws_table.max_row,
                                  max_col=ws_table.max_column, values_only=True):
        try:
            cur.execute("INSERT INTO det_data VALUES (?, ?, ?, ?)", row)
        except IntegrityError:
            print(f'{row[0]} is already in the table "det_data"')


def imp_det_list(cur, date: str):
    """
    Import data from a file
    """
    table_name = f'det_list_{date}.xlsx'
    data_path = abspath(join('.', table_name))

    wb = load_workbook(filename=data_path, data_only=True, read_only=True)
    s_name = list(wb.sheetnames)[0]
    ws_t = wb[s_name]
    imp_table_data(cur, ws_t)

    wb.close()


def file_name_re(name_re, file_list):
    """
    Extract file name according to regular expression
    """
    name_comp = re.compile(name_re)
    for name in file_list:
        name_m = name_comp.match(name)
        if name_m:
            return name_m.group()


def get_files_name(cur, path_dir):
    """
    Get file names according to the detail list (calculations)
    """
    data = cur.execute("SELECT ord_det FROM det_data").fetchall()
    list_calc = os.listdir('.' + path_dir)
    e_list = []     # detail list without calculation

    for d in data:
        name_re_calc = str(d[0] + '_r\d{2}_\w{3,4}.xlsx')
        file_list_calc = file_name_re(name_re_calc, list_calc)
        if file_list_calc:
            try:
                cur.execute("INSERT INTO calc_files VALUES (?, ?)", (d[0], file_list_calc))
            except IntegrityError:
                print(f'{d[0]} is already in the table "calc_files"')
        else:
            e_list.append(d[0])

    return e_list


def imp_det_calc(ord_det, ws, cur):
    """
    Get the data from the calculation table
    """
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=12, values_only=True):
        if row[0]:
            r1 = row[1]  # check if part number
            if r1 == 0 or r1 == 1:
                r1 = row[3]
            try:
                qty = int(row[2]) * int(row[9])
            except TypeError:
                qty = 0
            c_data = (ord_det, r1, row[2], row[3], row[8], row[9], row[10], qty)
            try:
                cur.execute('INSERT INTO calculation VALUES (?, ?, ?, ?, ?, ?, ?, ?)', c_data)
            except IntegrityError:  # repetition check
                c_data_2 = (row[2], qty, ord_det, r1)
                cur.execute("""UPDATE calculation
                                  SET part_qty = part_qty + ?,
                                      qty_per_det = qty_per_det + ?
                                WHERE ord_det = ? AND 
                                      part_num = ?
                            """, c_data_2)


def get_data_calc(cur, path_dir, s_n, e_list):
    """
    Get data from calculation files
    """
    f_list = cur.execute('SELECT DISTINCT * FROM calc_files').fetchall()

    for d in f_list:
        data_path = abspath(join('.' + path_dir, d[1]))
        wb = load_workbook(filename=data_path, data_only=True, read_only=True)
        s_name = list(wb.sheetnames)
        for sn in s_name:
            if sn == s_n:
                ws = wb[sn]
                imp_det_calc(d[0], ws, cur)
        wb.close()

    cur.execute("""SELECT DISTINCT cf.ord_det
                              FROM calc_files cf
                                   LEFT JOIN calculation ca ON cf.ord_det = ca.ord_det
                             WHERE ca.ord_det IS NULL
                """)
    empty = cur.fetchall()
    for d in empty:
        e_list.append(d[0])

    return e_list
