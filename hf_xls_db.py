import os
import sqlite3
from os.path import join, abspath
from sqlite3 import IntegrityError

from openpyxl import load_workbook


def get_hf_names(cur, path_dir):
    """
    Get file names according to the detail list (semi-finished product)
    """
    cur.row_factory = lambda cursor, row: row[0]
    ord_list = cur.execute("SELECT ord FROM det_data").fetchall()
    cur.row_factory = None

    list_hf_files = os.listdir('.' + path_dir)
    for f in list_hf_files:  # select the files that are in the order list
        d = f.removesuffix('.xlsx')
        if d in ord_list:
            try:
                cur.execute("INSERT INTO hf_files VALUES (?, ?)", (d, f))
            except IntegrityError:
                pass

    return cur.execute("SELECT * FROM hf_files").fetchall()


def imp_hf_data(cur, path_dir):
    """
    Get data from semi-finished products
    """
    list_hf_files_ = get_hf_names(cur, path_dir)

    for file in list_hf_files_:
        data_path = abspath(join('.' + path_dir, file[1]))
        wb = load_workbook(filename=data_path, data_only=True, read_only=True)
        s_name = list(wb.sheetnames)

        for sn in s_name:
            if sn == 'Аркуш1':
                ws = wb[sn]
                for row in ws.iter_rows(min_row=4, min_col=1, max_row=ws.max_row, max_col=5, values_only=True):
                    if row[0]:
                        r1 = str(row[0]) + '_' + str(row[1])
                        try:
                            cur.execute("INSERT INTO halffabricat VALUES (?, ?, ?)", (r1, row[2], row[4]))
                        except IntegrityError:
                            pass


def get_data_calc_hf(db, path_dir):
    """
    Replace the material with semi-finished product.
    If the part has a semi-finished product, we must replace the material with it
    from the directory of semi-finished products get file names according to the order list.
    """
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    imp_hf_data(cur, path_dir)

    cur.execute("""
                    UPDATE calculation
                       SET mat_type = 'halffabricat',
                           qty_per_det = (s.qty / calculation.part_qty) 
                      FROM (
                               SELECT c.ord_det,
                                      c.part_num,
                                      h.qty,
                                      c.part_qty
                                 FROM calculation c
                                      JOIN
                                      halffabricat h ON c.ord_det = h.ord_det AND 
                                                        c.part_num = h.part_num
                           ) AS s
                     WHERE calculation.ord_det = s.ord_det AND 
                           calculation.part_num = s.part_num
                """)

    conn.commit()
    conn.close()
